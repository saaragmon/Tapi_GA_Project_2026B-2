import os
import time
import random
from openpyxl import load_workbook

from Tapi_Project_Solver_Engine import (
    _xlsx_name,
    _read_input,
    _solve_fixed_order,
    _write_output
)


def _read_ga_params(input_file):
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    population_size = int(ws["A7"].value)
    elite_count = int(ws["A9"].value)
    mutation_probability = float(ws["A11"].value)
    time_limit = float(ws["A13"].value)

    if population_size % 2 != 0:
        raise ValueError("Population size must be an even number.")

    if elite_count % 2 != 0:
        raise ValueError("Number of solutions passed to next generation must be even.")

    if elite_count > int(0.1 * population_size):
        raise ValueError("Number of solutions passed cannot be more than 10% of population size.")

    if not (20 <= population_size <= 200):
        raise ValueError("Population size must be between 20 and 200.")

    if not (0 <= mutation_probability <= 1):
        raise ValueError("Mutation probability must be between 0 and 1.")

    if time_limit <= 0:
        raise ValueError("Time limit must be positive.")

    return population_size, elite_count, mutation_probability, time_limit


def _evaluate_order(order, jobs_by_id, cache):
    key = tuple(order)

    if key in cache:
        return cache[key]

    ordered_jobs = [jobs_by_id[j] for j in order]
    solution, value = _solve_fixed_order(ordered_jobs)

    result = {
        "order": list(order),
        "solution": solution,
        "value": value
    }

    cache[key] = result
    return result


def _ranking_probabilities(population):
    ranked = sorted(population, key=lambda x: x["value"])
    k = len(ranked)

    ranked_probs = []

    for rank, individual in enumerate(ranked, start=1):
        probability = 2 * (k - rank + 1) / (k * (k + 1))
        ranked_probs.append((individual, probability))

    return ranked_probs


def _select_by_probability(ranked_probs):
    r = random.random()
    cumulative = 0

    for individual, probability in ranked_probs:
        cumulative += probability
        if r <= cumulative:
            return individual

    return ranked_probs[-1][0]


def _order_crossover(parent1, parent2):
    n = len(parent1)
    start, end = sorted(random.sample(range(n), 2))

    child = [None] * n
    child[start:end + 1] = parent1[start:end + 1]

    missing_genes = [gene for gene in parent2 if gene not in child]

    idx = 0
    for i in range(n):
        if child[i] is None:
            child[i] = missing_genes[idx]
            idx += 1

    return child


def _mutate(order):
    mutated = list(order)
    i, j = random.sample(range(len(mutated)), 2)
    mutated[i], mutated[j] = mutated[j], mutated[i]
    return mutated


def Run_GA(Input_File_Name, Output_File_Na):
    start_time = time.time()

    input_file = _xlsx_name(Input_File_Name)
    output_file = _xlsx_name(Output_File_Na)

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file not found: {input_file}")

    jobs = _read_input(input_file)
    jobs_by_id = {job["j"]: job for job in jobs}
    job_ids = [job["j"] for job in jobs]

    population_size, elite_count, mutation_probability, time_limit = _read_ga_params(input_file)
    deadline = start_time + time_limit

    print("\n========== STARTING GENETIC ALGORITHM ==========")
    print(f"Population Size: {population_size}")
    print(f"Solutions Passed to Next Generation: {elite_count}")
    print(f"Mutation Probability: {mutation_probability}")
    print(f"Time Limit: {time_limit} seconds")
    print("===============================================\n")

    cache = {}

    original_solution = _evaluate_order(job_ids, jobs_by_id, cache)
    original_value = original_solution["value"]

    best_solution = original_solution
    best_generation = 0

    population = []

    for _ in range(population_size):
        if time.time() >= deadline:
            break

        random_order = random.sample(job_ids, len(job_ids))
        evaluated = _evaluate_order(random_order, jobs_by_id, cache)

        population.append(evaluated)

        if evaluated["value"] < best_solution["value"]:
            best_solution = evaluated
            best_generation = 0

    generation = 0
    last_printed_best = best_solution["value"]

    while time.time() < deadline:
        generation += 1

        print(f"Generation {generation}")

        ranked_probs = _ranking_probabilities(population)
        next_population = []

        for _ in range(elite_count):
            if time.time() >= deadline:
                break

            selected = _select_by_probability(ranked_probs)
            next_population.append(selected)

        while len(next_population) < population_size and time.time() < deadline:
            parent1 = _select_by_probability(ranked_probs)
            parent2 = _select_by_probability(ranked_probs)

            child1_order = _order_crossover(parent1["order"], parent2["order"])
            child2_order = _order_crossover(parent2["order"], parent1["order"])

            child1 = _evaluate_order(child1_order, jobs_by_id, cache)
            next_population.append(child1)

            if child1["value"] < best_solution["value"]:
                best_solution = child1
                best_generation = generation

            if len(next_population) < population_size and time.time() < deadline:
                child2 = _evaluate_order(child2_order, jobs_by_id, cache)
                next_population.append(child2)

                if child2["value"] < best_solution["value"]:
                    best_solution = child2
                    best_generation = generation

        mutated_population = []

        for individual in next_population:
            if time.time() >= deadline:
                mutated_population.append(individual)
                continue

            if random.random() < mutation_probability:
                mutated_order = _mutate(individual["order"])
                mutated_individual = _evaluate_order(mutated_order, jobs_by_id, cache)

                mutated_population.append(mutated_individual)

                if mutated_individual["value"] < best_solution["value"]:
                    best_solution = mutated_individual
                    best_generation = generation
            else:
                mutated_population.append(individual)

        population = mutated_population

        if generation % 5 == 0:
            print(f"Generation {generation} | Best value so far: {best_solution['value']}")

            if best_solution["value"] < last_printed_best:
                print(f"New best solution found in generation {best_generation}: {best_solution['value']}")
                last_printed_best = best_solution["value"]

    runtime = time.time() - start_time

    _write_output(
        output_file,
        best_solution["solution"],
        best_solution["value"],
        runtime
    )

    wb = load_workbook(output_file)
    ws = wb.active

    ws["B12"] = best_solution["value"]
    ws["B13"] = original_value
    ws["B14"] = generation
    ws["B15"] = round(runtime, 2)

    wb.save(output_file)

    print("\n========== FINISHED ==========")
    print(f"Best value found: {best_solution['value']}")
    print(f"Original order value: {original_value}")
    print(f"Generations created: {generation}")
    print(f"Runtime: {round(runtime, 2)} seconds")