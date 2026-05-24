import os
import shutil
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pulp


def _xlsx_name(file_name):
    return file_name if file_name.lower().endswith(".xlsx") else file_name + ".xlsx"


def _read_input(input_file):
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    jobs = []
    col = 2
    while ws.cell(2, col).value not in (None, ""):
        jobs.append({
            "j": int(ws.cell(1, col).value),
            "tj": float(ws.cell(2, col).value),
            "dj": float(ws.cell(3, col).value),
            "vj": float(ws.cell(4, col).value),
            "wj": float(ws.cell(5, col).value),
        })
        col += 1

    return jobs


def _solve_fixed_order(jobs):
    prob = pulp.LpProblem("JIT_Scheduling_Fixed_Order", pulp.LpMinimize)

    S, C, E, T = {}, {}, {}, {}

    for job in jobs:
        j = job["j"]
        S[j] = pulp.LpVariable(f"S_{j}", lowBound=0)
        C[j] = pulp.LpVariable(f"C_{j}", lowBound=0)
        E[j] = pulp.LpVariable(f"E_{j}", lowBound=0)
        T[j] = pulp.LpVariable(f"T_{j}", lowBound=0)

    prob += pulp.lpSum(job["vj"] * E[job["j"]] + job["wj"] * T[job["j"]] for job in jobs)

    for idx, job in enumerate(jobs):
        j = job["j"]
        tj = job["tj"]
        dj = job["dj"]

        prob += C[j] == S[j] + tj
        prob += E[j] >= dj - C[j]
        prob += T[j] >= C[j] - dj

        if idx > 0:
            prev_j = jobs[idx - 1]["j"]
            prob += S[j] >= C[prev_j]

    solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=120)
    status = prob.solve(solver)

    if pulp.LpStatus[status] not in ("Optimal", "Feasible"):
        raise Exception(f"Solver failed. Status: {pulp.LpStatus[status]}")

    result = []
    total_penalty = 0

    for job in jobs:
        j = job["j"]
        sj = float(pulp.value(S[j]))
        cj = float(pulp.value(C[j]))
        ej = max(0, job["dj"] - cj)
        tj_late = max(0, cj - job["dj"])
        pj = job["vj"] * ej + job["wj"] * tj_late
        total_penalty += pj

        result.append({
            **job,
            "Sj": sj,
            "Cj": cj,
            "Ej": ej,
            "Tj": tj_late,
            "Pj": pj,
        })

    return result, total_penalty


def _build_segments(solution):
    segments = []
    current_time = 0

    for job in solution:
        if job["Sj"] > current_time:
            segments.append({
                "type": "IDLE",
                "label": f"IDLE_{current_time:g}-{job['Sj']:g}",
                "start": current_time,
                "end": job["Sj"],
            })

        segments.append({
            "type": "JOB",
            "label": f"j{int(job['j'])}",
            "start": job["Sj"],
            "end": job["Cj"],
            "job": job,
        })

        current_time = job["Cj"]

    return segments


def _write_output(output_file, solution, total_penalty, runtime):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    segments = _build_segments(solution)

    headers = ["J", "tj", "dj", "vj", "wj", "Sj", "Cj", "Ej", "Tj", "Pj"]
    for r, header in enumerate(headers, start=1):
        ws.cell(r, 1).value = header

    for c, seg in enumerate(segments, start=2):
        if seg["type"] == "IDLE":
            ws.cell(1, c).value = seg["label"]
            ws.cell(2, c).value = "-"
            ws.cell(3, c).value = "-"
            ws.cell(4, c).value = "-"
            ws.cell(5, c).value = "-"
            ws.cell(6, c).value = seg["start"]
            ws.cell(7, c).value = seg["end"]
            ws.cell(8, c).value = "-"
            ws.cell(9, c).value = "-"
            ws.cell(10, c).value = "-"
        else:
            job = seg["job"]
            ws.cell(1, c).value = f"j{int(job['j'])}"
            ws.cell(2, c).value = job["tj"]
            ws.cell(3, c).value = job["dj"]
            ws.cell(4, c).value = job["vj"]
            ws.cell(5, c).value = job["wj"]
            ws.cell(6, c).value = job["Sj"]
            ws.cell(7, c).value = job["Cj"]
            ws.cell(8, c).value = job["Ej"]
            ws.cell(9, c).value = job["Tj"]
            ws.cell(10, c).value = job["Pj"]

    ws["A12"] = "Total Penalty"
    ws["B12"] = total_penalty
    ws["A13"] = "Original Sol"
    ws["B13"] = total_penalty
    ws["A14"] = "Total Gen Created"
    ws["B14"] = "-"
    ws["A15"] = "Run time"
    ws["B15"] = round(runtime, 4)

    ws["A17"] = "Gantt:"

    job_fill = PatternFill("solid", fgColor="9AD0F5")
    idle_fill = PatternFill("solid", fgColor="D9D9D9")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 11):
        ws.cell(row, 1).font = bold_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 1).font = white_font

    for col in range(1, len(segments) + 2):
        ws.cell(1, col).fill = header_fill
        ws.cell(1, col).font = white_font
        for row in range(1, 16):
            ws.cell(row, col).border = border
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

    for c, seg in enumerate(segments, start=2):
        duration = seg["end"] - seg["start"]
        ws.cell(17, c).value = f"{seg['label']}\n({seg['start']:g} - {seg['end']:g})"
        ws.cell(17, c).fill = idle_fill if seg["type"] == "IDLE" else job_fill
        ws.cell(17, c).border = border
        ws.cell(17, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(25, duration * 1.5))

    ws["A17"].font = bold_font
    ws["A17"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[17].height = 45
    ws.column_dimensions["A"].width = 18

    for row in [12, 13, 14, 15]:
        ws.cell(row, 1).font = bold_font
        ws.cell(row, 1).fill = header_fill
        ws.cell(row, 1).font = white_font
        ws.cell(row, 2).alignment = Alignment(horizontal="center")

    wb.save(output_file)


def Run_Solver(Input_File_Name, Output_File_Na):
    start_time = time.time()

    input_file = _xlsx_name(Input_File_Name)
    output_file = _xlsx_name(Output_File_Na)

    if not os.path.exists(input_file):
        raise FileNotFoundError(f"Input file was not found: {input_file}")

    jobs = _read_input(input_file)
    solution, total_penalty = _solve_fixed_order(jobs)

    runtime = time.time() - start_time
    _write_output(output_file, solution, total_penalty, runtime)

    print(f"Solver finished successfully.")
    print(f"Output file created: {output_file}")
    print(f"Total penalty: {total_penalty}")